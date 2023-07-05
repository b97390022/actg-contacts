from src.db import RedisClient
from watchdog.observers.polling import PollingObserver
from watchdog.events import FileSystemEventHandler
from loguru import logger
from src.utils import get_file_path
import re
from openpyxl import load_workbook
import unicodedata
from pathlib import Path

redis_client = RedisClient()


class MyEventHandler(FileSystemEventHandler):
    def __init__(self) -> None:
        super().__init__()

    def on_modified(self, event):
        logger.info(f"File modified: {event.src_path}, {event.event_type}")
        # Define the actions to be performed when a file is modified


class ContactService:
    def __init__(self) -> None:
        self.src_path: str = r"/mnt/Public/公司通訊錄及座位表"
        self.contact_file: str = "*通訊錄*.xlsx"
        self.contact_file_path: str = get_file_path(self.src_path, self.contact_file)
        self.col_index: dict = {
            "idx_phone": 0,
            "idx_office": 1,
            "idx_department": 2,
            "idx_chinese_name": 4,
            "idx_english_name": 5,
            "idx_email": 6,
        }
        self.attributes: dict = {
            "name": "姓名",
            "department": "部門",
            "email": "信箱",
            "phone": "分機號碼",
        }
        self.default_no_data_string: str = "無資料"
        self.split_line: str = f"{'-'*40}\n"

    def get_split_line_length(self):
        return len(self.split_line)

    def append_split_line(self, text_string: str):
        text_string += self.split_line
        return text_string

    def _get_colname_index(self, sheet):
        # =============================================================================
        #     lookup the correct column index by impossible-to-change content
        # =============================================================================
        for row in sheet.iter_rows():
            list_cell = list(map(lambda x: x.value, row))
            if 1100 in list_cell:
                idx_phone = list_cell.index(1100)
            if "分子檢驗處" in list_cell:
                idx_office = list_cell.index("分子檢驗處")
            if "次世代定序部" in list_cell:
                idx_department = list_cell.index("次世代定序部")
            if "陳華鍵" in list_cell:
                idx_chinese_name = list_cell.index("陳華鍵")
            if "Hua-Chien Chen" in list_cell:
                idx_english_name = list_cell.index("Hua-Chien Chen")
            if "hcchen@actgenomics.com" in list_cell:
                idx_email = list_cell.index("hcchen@actgenomics.com")
        return (
            idx_phone,
            idx_office,
            idx_department,
            idx_chinese_name,
            idx_english_name,
            idx_email,
        )

    def _parse_contact(self, sheet):
        return dict(
            zip(
                (
                    "idx_phone",
                    "idx_office",
                    "idx_department",
                    "idx_chinese_name",
                    "idx_english_name",
                    "idx_email",
                ),
                self._get_colname_index(sheet),
            )
        )

    def _iter_contact_data(self, sheet):
        """
        generator of excel data
        """
        start = finish = False
        for row in sheet.iter_rows():
            list_row = list(row)
            (
                phone,
                office,
                department,
                chinese_name,
                english_name,
                email,
            ) = (list_row[x].value for x in (self.col_index.values()))

            if phone in [1100, 1500, 1200]:
                start = True
            if start and email is None:
                finish = True

            if start and not finish:
                yield (
                    phone if phone else self.default_no_data_string,
                    office if office else self.default_no_data_string,
                    unicodedata.normalize("NFKC", department)
                    if department
                    else self.default_no_data_string,
                    unicodedata.normalize("NFKC", chinese_name.replace(" ", ""))
                    if chinese_name
                    else self.default_no_data_string,
                    unicodedata.normalize("NFKC", english_name)
                    if english_name
                    else self.default_no_data_string,
                    unicodedata.normalize("NFKC", email)
                    if email
                    else self.default_no_data_string,
                )
            elif finish:
                break

    def _key_string_modify(self, key_string):
        return "@%s@" % "@".join(list(key_string.strip()))

    def _get_department_all(self, sheet):
        return [
            department
            for (
                phone,
                office,
                department,
                chinese_name,
                english_name,
                email,
            ) in self._iter_contact_data(sheet)
        ]

    def make_dict(self, sheet_name: str):
        sheet = load_workbook(self.contact_file_path)[sheet_name]
        self.col_index = self._parse_contact(sheet)
        excel_file_name = Path(self.contact_file_path).name
        # Make dictionary
        dict_mapping_data = {"*version": excel_file_name, "*ver": excel_file_name}

        for (
            phone,
            office,
            department,
            chinese_name,
            english_name,
            email,
        ) in self._iter_contact_data(sheet):
            key_ = f"{english_name} ({chinese_name})"
            dict_mapping_data[key_] = {
                "name": key_,
                "department": f"{department},{office}",
                "email": str(email),
                "phone": str(phone),
            }
        return dict_mapping_data

    def serialize_to_split_list(self, elements: list[dict], max_length: int):
        split_line_length = self.get_split_line_length()
        m_length = max_length - 2 * split_line_length
        l = len(elements)
        result_list = []
        result = ""
        result = self.append_split_line(result)  # start
        curr_length = 0
        curr_length += split_line_length

        while elements:
            element = elements.pop()
            t_str = ""
            for k, v in self.attributes.items():
                t_str += f"**{v}**: {element.get(k) or '無資料'}\n"
            t_length = len(t_str)
            if curr_length + t_length < m_length:
                result += t_str
                result = self.append_split_line(result)
                curr_length += t_length
                curr_length += split_line_length
            else:
                result_list.append(result)
                result = ""
                result = self.append_split_line(result)  # start
                curr_length = 0
                curr_length += split_line_length

        if l == 0:
            result = "找不到聯絡人。\n"
        result_list.append(result)
        return result_list

    async def search(self, search_string: str, max_length: int):
        contact_dict = self.make_dict("ACTG")
        results = []
        for contact_name in contact_dict:
            r = re.findall(r"\(.*\)", search_string, re.IGNORECASE)
            if r and re.search(r[0], contact_name, re.IGNORECASE):
                results.append(contact_dict[contact_name])
            elif re.search(search_string, contact_name, re.IGNORECASE):
                results.append(contact_dict[contact_name])
        return self.serialize_to_split_list(results, max_length)


if __name__ == "__main__":
    import asyncio

    contact_service = ContactService()

    async def main():
        r = await contact_service.search("Jian Siao Yu (簡孝羽)", 2000)
        print(r)

    asyncio.run(main())
    # logger.info(contact_service.contact_file_path)
    # observer = PollingObserver()
    # observer.schedule(MyEventHandler(), path=contact_service.src_path, recursive=False)
    # observer.start()
    # try:
    #     while observer.is_alive():
    #         observer.join(1)
    # finally:
    #     observer.stop()
    #     observer.join()
