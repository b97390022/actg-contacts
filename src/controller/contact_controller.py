from __future__ import annotations

import re
import unicodedata
import uuid
from pathlib import Path

import xlwings as xw
from openpyxl import load_workbook

from src.component.image import draw_red_circle, read_image
from src.config import config
from src.utils import get_file_path


class ContactController:
    def __init__(self) -> None:
        self.src_path: str = config.env_config.contact_src_path
        self.contact_file: str = "*通訊錄*.xlsx"
        self.col_index: dict = {
            "idx_phone": 0,
            "idx_office": 1,
            "idx_department": 2,
            "idx_chinese_name": 4,
            "idx_english_name": 5,
            "idx_email": 6,
        }
        self.attributes: dict = {
            "name": "姓名",
            "department": "部門",
            "email": "信箱",
            "phone": "分機號碼",
        }
        self.default_no_data_string: str = "無資料"
        self.split_line: str = f"{'-'*40}\n"


    def append_split_line(self, text_string: str):
        text_string += self.split_line
        return text_string

    def _get_colname_index(self, sheet):
        # ============================================================================= #
        #     lookup the correct column index by impossible-to-change content           #
        # ============================================================================= #
        for row in sheet.iter_rows():
            list_cell = [x.value for x in row]
            if 1100 in list_cell:
                idx_phone = list_cell.index(1100)
            if "分子檢驗處" in list_cell:
                idx_office = list_cell.index("分子檢驗處")
            if "次世代定序部" in list_cell:
                idx_department = list_cell.index("次世代定序部")
            if "陳華鍵" in list_cell:
                idx_chinese_name = list_cell.index("陳華鍵")
            if "Hua-Chien Chen" in list_cell:
                idx_english_name = list_cell.index("Hua-Chien Chen")
            if "hcchen@actgenomics.com" in list_cell:
                idx_email = list_cell.index("hcchen@actgenomics.com")
        return (
            idx_phone,
            idx_office,
            idx_department,
            idx_chinese_name,
            idx_english_name,
            idx_email,
        )

    def _parse_contact(self, sheet):
        return dict(
            zip(
                (
                    "idx_phone",
                    "idx_office",
                    "idx_department",
                    "idx_chinese_name",
                    "idx_english_name",
                    "idx_email",
                ),
                self._get_colname_index(sheet),
            )
        )

    def _iter_contact_data(self, sheet):
        """
        generator of excel data
        """
        start = finish = False
        for row in sheet.iter_rows():
            list_row = list(row)
            (
                phone,
                office,
                department,
                chinese_name,
                english_name,
                email,
            ) = (list_row[x].value for x in (self.col_index.values()))

            if phone in {1100, 1500, 1200}:
                start = True
            if start and email is None:
                finish = True

            if start and not finish:
                yield (
                    phone if phone else self.default_no_data_string,
                    office if office else self.default_no_data_string,
                    unicodedata.normalize("NFKC", department)
                    if department
                    else self.default_no_data_string,
                    unicodedata.normalize("NFKC", chinese_name.replace(" ", ""))
                    if chinese_name
                    else self.default_no_data_string,
                    unicodedata.normalize("NFKC", english_name)
                    if english_name
                    else self.default_no_data_string,
                    unicodedata.normalize("NFKC", email)
                    if email
                    else self.default_no_data_string,
                )
            elif finish:
                break

    def _key_string_modify(self, key_string):
        return f'@{"@".join(list(key_string.strip()))}@'

    def _get_department_all(self, sheet):
        return [
            department
            for (
                phone,
                office,
                department,
                chinese_name,
                english_name,
                email,
            ) in self._iter_contact_data(sheet)
        ]

    def get_split_line_length(self):
        return len(self.split_line)

    def make_dict(self, sheet_name: str):
        contact_file_path = get_file_path(self.src_path, self.contact_file)
        sheet = load_workbook(contact_file_path)[sheet_name]
        self.col_index = self._parse_contact(sheet)
        excel_file_name = Path(contact_file_path).name
        # Make dictionary
        dict_mapping_data = {"*version": excel_file_name, "*ver": excel_file_name}

        for (
            phone,
            office,
            department,
            chinese_name,
            english_name,
            email,
        ) in self._iter_contact_data(sheet):
            key_ = f"{english_name} ({chinese_name})"
            dict_mapping_data[key_] = {
                "name": key_,
                "department": f"{department},{office}",
                "email": str(email),
                "phone": str(phone),
            }
        return dict_mapping_data

    def serialize_to_split_list(self, elements: list[dict], max_length: int):
        split_line_length = self.get_split_line_length()
        m_length = max_length - 2 * split_line_length
        l = len(elements)
        result_list = []
        result = ""
        result = self.append_split_line(result)  # start
        curr_length = 0
        curr_length += split_line_length

        while elements:
            element = elements.pop()
            t_str = ""
            for k, v in self.attributes.items():
                t_str += f"**{v}**: {element.get(k) or '無資料'}\n"
            t_length = len(t_str)
            if curr_length + t_length < m_length:
                result += t_str
                result = self.append_split_line(result)
                curr_length += t_length
                curr_length += split_line_length
            else:
                result_list.append(result)
                result = ""
                result = self.append_split_line(result)  # start
                curr_length = 0
                curr_length += split_line_length

        if l == 0:
            result = "找不到聯絡人。\n"
        result_list.append(result)
        return result_list

    async def search(self, search_string: str, max_length: int):
        contact_dict = self.make_dict("ACTG")
        results = []
        for contact_name, contact_value in contact_dict.items():
            r = re.findall(r"\(.*\)", search_string, re.IGNORECASE)
            if r and re.search(r[0], contact_name, re.IGNORECASE):
                results.append(contact_value)
            elif re.search(search_string, contact_name, re.IGNORECASE):
                results.append(contact_value)
        return self.serialize_to_split_list(results, max_length)


class SeatingChartController:
    def __init__(self) -> None:
        self.src_path: str = config.env_config.contact_src_path
        self.seating_chart_file: str = "*座位圖*.xlsx"
        self.circle_offset: int = 100

    def _formalize(self, input_string: str):
        input_string = unicodedata.normalize("NFKC", input_string)
        input_string = input_string.lower()
        input_string = re.sub(r"[- ]+", "", input_string)
        return input_string

    def _iter_seating_chart_data(
        self,
        sheet: xw.Sheet,
        border: tuple[int, int] | None = None,
        x: float = 0.0,
        y: float = 0.0
    ):
        """
        generator of excel data
        """
        if not border:
            lrow, lcol = self._get_border(sheet)
        else:
            lrow, lcol = border

        for i in range(lrow):
            y=0.0
            for j in range(lcol):
                cell = sheet.cells(i, j)
                if cell.value and cell.value != '':
                    yield (cell, y, x)
                y+=cell.width
            x+=sheet.cells(i, 1).height

    def _get_all_sheets(self, wb: xw.Book) -> list[str]:
            return wb.sheet_names

    def _get_border(self, sheet: xw.Sheet):
        return (sheet.used_range.last_cell.row, sheet.used_range.last_cell.column)

    def draw_circle_on_png(self, l: list[tuple[str, int, int]]) -> list:

        for (file_name, x, y) in l:
            image = read_image(file_name)
            xy =  (x - self.circle_offset/2, y - self.circle_offset/2, x + self.circle_offset/2, y + self.circle_offset/2)
            image = draw_red_circle(
                image=image,
                xy=xy,
            )
            image.save(file_name)


    async def search(self, search_string: str):
        search_string = self._formalize(search_string)
        seating_chart_file_path = get_file_path(self.src_path, self.seating_chart_file)
        images_path = []
        with xw.Book(seating_chart_file_path, read_only=True) as wb:

            for sheet_name in  self._get_all_sheets(wb):
                sheet: xw.Sheet = wb.sheets[sheet_name]
                border = self._get_border(sheet)

                l=[]
                for (cell, x, y) in self._iter_seating_chart_data(sheet, border):
                    value = self._formalize(cell.value)
                    if value != "" and (search_string in value or value in search_string):
                        file_name = f'images/{uuid.uuid4()}.png'
                        l.append((file_name, x, y))
                        sheet.range((1,1), border).to_png(file_name)
                        images_path.append(file_name)
                        break
                self.draw_circle_on_png(l)
                if images_path:
                    break

        return images_path

if __name__ == "__main__":
    import asyncio

    # contact_controller = ContactController()
    seating_chart_controller = SeatingChartController()
    # r = seating_chart_controller.get_all_sheets()
    async def main():
        await seating_chart_controller.search("陳柏劭")


    asyncio.run(main())
